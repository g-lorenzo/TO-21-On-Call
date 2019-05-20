from openpyxl   import load_workbook
import os
import bs4

#Get the height, aka number of rows, in the sheet of a workbook given the string for calculated dimensions
# def getsheetheight(attr_list):
#     dim = re.findall(r"\d+",str) #Get numeric values from dimensions string
#     hgt = int(dim[1]) - 1 #Subtract 1 because Python starts counting at 1?
#     return(hgt)

attr1_names = ['Route - ', 'Mile Post - ', 'Description - ', 'Critical - ', 'Microwave Receiver Required for Video: ']
attr2_names = ['Project Name - ', 'P.I. # - ', 'Federal Project # - ']

"""
attr_text is the following list of strings:
[
ro = Route Name
mp = Mile Post
de = Description
cr = Critical Status
mr = Microwave Receiver Required for Video
prj_n = Project Name
pi = P.I. #
fp = Federal Project Number
]

"""

def createCDataText(attr_text):
    #Typical attribute names
    attr1_names = ['Route - ', 'Mile Post - ', 'Description - ', 'Critical - ', 'Microwave Receiver Required for Video: ']
    attr2_names = ['Project Name - ', 'P.I. # - ', 'Federal Project # - ']

    #Build an empty HTML container and add create a paragraph element for the container
    soup = bs4.BeautifulSoup(features="html.parser")
    paragraph = bs4.Tag(soup, name="p")

    # SECTION 1: Location Info

    h1_tag = soup.new_tag(name="font"); h1_tag['size'] = "4" # Location Info, size 4
    h1_str = soup.new_tag(name="strong") # Make it bold
    h1_str.string = "Location Info"

    br = soup.new_tag('br') # Add a break line

    h1_tag.append(h1_str) # Append bold text to header string
    paragraph.append(h1_tag) # Append header string to paragraph
    paragraph.append(br)

    attrs1 = attr_text[0:5]

    for attr in range(len(attr1_names)): # Iterate over attributes in attr1_names

        new_br = soup.new_tag('br')

        new_strong  =   bs4.Tag(soup, name="strong") # Make the attribute name bold
        new_strong.string = attr1_names[attr]

        new_text = bs4.Tag(soup, name="font"); new_text['size'] = "3"
        new_text.string = attrs1[attr]

        new_attr = bs4.Tag(soup, name="font"); new_attr['size'] = "3"

        new_attr.append(new_strong)
        new_attr.append(new_text)
        paragraph.append(new_attr)
        paragraph.append(new_br) # Add a break line

    br = soup.new_tag('br'); paragraph.append(br)

    # SECTION 2: Project Info

    attrs2 = attr_text[5:]

    h2_tag = soup.new_tag(name="font"); h2_tag['size'] = "4" # Location Info, size 4
    h2_str = soup.new_tag(name="strong") # Make it bold
    h2_str.string = "Project Info"

    br = soup.new_tag('br')

    h2_tag.append(h2_str) # Append bold text to header string
    paragraph.append(h2_tag) # Append header string to paragraph
    paragraph.append(br)

    for attr in range(len(attr2_names)):    #Iterate over attributes in attr2_names

        new_br = soup.new_tag('br')

        new_strong  =   bs4.Tag(soup, name="strong") # Make the attribute name bold
        new_strong.string = attr2_names[attr]

        new_text = bs4.Tag(soup, name="font"); new_text['size'] = "3"
        new_text.string = attrs2[attr]

        new_attr = bs4.Tag(soup, name="font"); new_attr['size'] = "3"

        new_attr.append(new_strong)
        new_attr.append(new_text)
        paragraph.append(new_attr)
        paragraph.append(new_br)

    soup.append(paragraph) # Append everything from above to the soup container

    return(str(soup))

def getExcelText(excelfile, sheetname=str, maxentries=0):

    wb = load_workbook(excelfile, read_only=True)

    if sheetname:
        parse_sheet = wb[sheetname]

    return('')

#[loc, desc, fail, bill.s, bill.f, bill.t, status, route, mp, gdotprojdesc]



if __name__ == '__main__':

    attr_index = [1, 1, 1, 0, 0, 0, 0, 1, 1, 1]

    datafolder = r'..\\data'
    fname = datafolder + r'\\' + 'NWC-SMC.xlsx'
    HTMLname = datafolder + r'\\' + 'XMLDataSample.html'

    wb = load_workbook(fname, read_only=True)

    sheet = wb['NWC']
    sheetDim = sheet.calculate_dimension()

    #Iterate over rows and take attributes based on attr_index list
    for r in range(1,10): #(sheetHeight):
        attr_values = []
        for a in range(len(attr_index)):
            if attr_index[a] == 1:
                attr = sheet.cell(row = r+1, column = a+1)
                attr_values.append(attr.value)

        # Use attributes and 'text' list to build soup
        text2 = ['', attr_values[4], attr_values[1], '', '', attr_values[5], '', '']




