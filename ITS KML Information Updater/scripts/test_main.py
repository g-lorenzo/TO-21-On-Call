import time

import ITSKMLParser as IKP

#from typicalFunctions import printProgressBar
from openpyxl import load_workbook

fname = r'..\data\serco databases\20190506_1734.xlsx'

wb = load_workbook(fname,read_only=True)
ws = wb['List of Locations']
num_rows = ws.max_row

del wb, ws

not_updated = []
updateFromKMLFlag = True # Set this to true when you want to potentially overwrite what is in existing KML with Excel
customFlags = [True]*9 #[ lat, lon, mp, proj, piNum, locat, dType, desc, fpi]
addDescriptionFlag = False #Allow user to add description for device if none present
added_devices = 0
updated_devices = 0
cant_add = 0
start = 0
end = num_rows-1
prog = start

# printProgressBar(0, ws.max_row, "Progress: ", "Complete!", 1)
ti = time.time()
for r in range(end):

    wb = load_workbook(fname, read_only=True)
    ws = wb['List of Locations']
    row = ws[r+2]
    del wb, ws

    if prog == start: #First iteration to create new file
        kmlFile = r'..\data\All Devices v6.0\doc.kml'
        newKml = kmlFile.replace('.kml', 'v20190510.kml')
    else: #Remaining iterations use the created file only
        kmlFile = newKml = r'..\data\All Devices v6.0\doc.kml'.replace('.kml', 'v20190510.kml')

    prog += 1 #Maintain count of the number of devices processed

    devName =       row[0].value
    devDesc =       row[1].value
    devFailCode =   row[2].value
    devRoute =      row[4].value
    devMP =         row[5].value
    devPrj =        ''
    devFol =        devFailCode

    if devFol == 'CCTV':
        devFol = 'CAM'
    elif devFol == 'CCTV-IR':
        devFol = 'CAM-IR'
    elif devFol == 'GLSS':
        devFol = 'GATE'
    elif devDesc == '(null)':
        #devDesc = input('Device ID {} has no description. Please enter one:'.format(devName));
        devDesc = ''

    try:
        devLat = float(row[13].value)
        devLon = float(row[14].value)
    except:
        continue
    del row

    try:
        # 1. Create a dummy device using the name and folder given from the excel sheet
        dummyDevice = IKP.Device(devFol, devName)

        # 1A. If data is being taken from the excel spreadsheet, update any relevant dummy device attributes [newDevice]
        dummyDevice.folder = devFol
        dummyDevice.desc = "{} - {}".format(devFailCode, devRoute)
        dummyDevice.dType = devFailCode
        dummyDevice.locat = devDesc
        dummyDevice.mp = devMP
        dummyDevice.project = devPrj
        dummyDevice.lat = devLat
        dummyDevice.lon = devLon

        del(devFol, devFailCode, devRoute, devDesc, devMP, devPrj, devLat, devLon)

        # 2. Given a name, check to see if it exists within the KML using the self.checkForDevice method
        if dummyDevice.checkForDevice(kmlFile): # 3. If the device already exists in the file

            # 3A. Retrieve the device attributes in the KML file using getExistingDeviceData function
            kmlDevice = IKP.getExistingDeviceData(kmlFile,dummyDevice)

            # If the kmlDevice is different from dummyDevice
            if kmlDevice != dummyDevice:
                # 3B. Update the dummy device using updateDummyDevice function (given updateFromKMLFlag == True)
                if updateFromKMLFlag:
                    dummyDevice.updateWithElement(kmlDevice)
                    updatedDevice = dummyDevice; del dummyDevice, kmlDevice

                else: # If updateFromKMLFlag == False
                    IKP.compareDevices(kmlDevice, dummyDevice, True)
                    updatedDevice = IKP.customUpdate(kmlDevice, dummyDevice, customFlags)
                    del dummyDevice, kmlDevice
            else:
                updatedDevice = kmlDevice; del kmlDevice, dummyDevice
            t1 = time.time()
            IKP.addDeviceToKML(updatedDevice, kmlFile, newKml, allowWrite=True, replace=True)
            t2 = time.time()
            print("UPDATING Device: {} -- Time Required (s): {}".format(updatedDevice.name, round(t2-t1,3)))
            del updatedDevice, t1, t2
            updated_devices += 1

            # 3C. Update device in KML and save
        else: # 4. If the device does NOT exist in the KML file:
            try:
                t1 = time.time()
                IKP.addDeviceToKML(dummyDevice, kmlFile, newKml, allowWrite=True, replace=False)
                t2 = time.time()
                print("ADDING Device: {} -- Time Required (s): {}".format(dummyDevice.name, round(t2-t1,3)))
                added_devices += 1
            except:
                print("Unable to add device {} to KM file".format(dummyDevice.name))
                cant_add += 1

    except: # When encountering any errors, print the device name for debugging
        not_updated.append(devName)

    if prog % 100 == 0:
        print("Completed {} of {} entries ({}%)".format(prog, num_rows, round(100 * prog / num_rows, 2)))

    # printProgressBar(prog, ws.max_row, "Progress: ", "Complete!", 1)
    # if prog % 25 == 0:
    #     print("Completed {} of {} entries ({}%)".format(prog, ws.max_row, round(100*prog/ws.max_row,2)))
tf = time.time()
print("---------- SUMMARY --------------------")
print("Total Devices Added: {}".format(added_devices))
print("Total Devices Updated: {}".format(updated_devices))
print("Total Devices Not Added/Updated: {}".format(cant_add))
print("Total Time: {}".format(round(tf-ti,3)))
print('---------------------------------------')


"""
File Name: test_main.py

Description: Test file for utilizing the functions in ITSKMLParser.py. This is the file that is manipulated to read
excel spreadsheets and pull the data for manipulation of a KML file.

Order of Operations:
    For any given device from an excel spreadsheet, you should follow the orders below to create or update that device
    in a KML file
    
    1. Create a dummy device given the device name (i.e. 'GDOT-CAM-92') and the folder (e.g. 'CAM', 'VSLS, etc.)
        1A. If data is being taken from the excel spreadsheet, update any relevant dummy device attributes [dummyDevice]
    2. Given the name (e.g. 'GDOT-CAM-92') check to see if it exists within the KML using the self.checkForDevice method
    3. If the device exists in the KML file:
        3A. Retrieve the device attributes in the KML file using getExistingDeviceData function
            i. If the excel spreadsheet has data to update, create a device for what is in the kml [kmlDevice]
        3B. Update the dummy device using updateDummyDevice function (given updateFromKMLFlag == 1)
            i. If the excel spreadsheet has data to update, compare newDevice with kmlDevice
                a. If there are differences, have the option to notify user with the following format:
                
                '''
                Device: 'GDOT-CAM-92' - Folder: 'CAM'
                    Latitude: {KML: 33.246925, EXCEL: 33.246952}
                    Longitude: {KML: -84.245236, EXCEL: -84.2452}
                    Project: {KML:  , EXCEL: 'NWC'}
                    ...
    4. If the device does NOT exist in the KML file:
        4A. Add the device to the KML using addDevicetoKML function
        
"""